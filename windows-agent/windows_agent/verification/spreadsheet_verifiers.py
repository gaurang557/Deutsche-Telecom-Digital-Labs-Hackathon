"""Verifiers for the `spreadsheet.*` actions (Milestone 4).

CORE PRINCIPLE (repeated because it matters)
--------------------------------------------
"The executor returned success" is NOT proof. The one modifying spreadsheet
action, `spreadsheet.write_cell`, gets an independent verifier that RE-OPENS the
workbook and RE-READS the target cell, then PASSes iff the observed value equals
the value we intended to write:

  * spreadsheet.write_cell → reload the workbook + re-read the cell; the cell's
                             value equals the intended written value.

Read-only actions (list_sheets / dimensions / read_cell / read_range) have no
verifier — the VerificationRegistry returns SKIPPED for them, which is correct:
there is no state change to confirm.

INDEPENDENCE
------------
The verifier opens the workbook itself (it does not trust the executor's live
handle) and takes the intended path, sheet, cell, and value only from the
authorized Action. ExecutorResult evidence is never accepted as the expectation,
so a buggy/lying executor cannot redefine what counts as success.

NUMBER-vs-STRING COMPARISON
---------------------------
Values are compared type-awarely: two numbers compare numerically (so writing
the int `42` and re-reading the float `42.0` still PASSes), while everything
else compares by normalised equality (e.g. `"42"` the string does NOT equal
`42` the number). Dates are normalised to ISO strings on both sides (matching
the executor), so they compare as strings.

The re-read uses data_only=False: we wrote a *literal* value, so there is no
cached-formula ambiguity to worry about (see `executors/spreadsheet_ops.py`).
Verification runs off the event loop via `asyncio.to_thread` (blocking I/O).
"""

from __future__ import annotations

import asyncio
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..contracts import Action, ExecutorResult, VerificationResult, VerificationStatus
from ..executors.spreadsheet_ops import _normalize_value
from .base import Verifier


def _passed(method: str, expected: Any, observed: Any, message: str = "") -> VerificationResult:
    return VerificationResult(
        status=VerificationStatus.PASSED, method=method, expected=expected, observed=observed, message=message
    )


def _failed(method: str, expected: Any, observed: Any, message: str) -> VerificationResult:
    return VerificationResult(
        status=VerificationStatus.FAILED, method=method, expected=expected, observed=observed, message=message
    )


def _values_match(expected: Any, observed: Any) -> bool:
    """Type-aware equality: numbers compare numerically, else normalised ==."""
    exp_num = isinstance(expected, (int, float)) and not isinstance(expected, bool)
    obs_num = isinstance(observed, (int, float)) and not isinstance(observed, bool)
    if exp_num and obs_num:
        return float(expected) == float(observed)
    return expected == observed


class SpreadsheetWriteCellVerifier(Verifier):
    """Confirm the workbook cell now holds the value we intended to write."""

    async def verify(self, action: Action, result: ExecutorResult, context: Any = None) -> VerificationResult:
        return await asyncio.to_thread(self._check, action, result)

    @staticmethod
    def _check(action: Action, result: ExecutorResult) -> VerificationResult:
        method = "reload workbook and re-read cell"
        params = action.parameters or {}

        path_str = action.target
        cell_ref = params.get("cell")
        sheet_name = params.get("sheet")
        expected = _normalize_value(params.get("value"))

        if not path_str or not isinstance(cell_ref, str):
            return _failed(method, "path+cell", None, "Missing path/cell for verification")
        path = Path(path_str)
        if not path.exists():
            return _failed(method, f"exists({path})", False, f"Workbook missing after write: {path}")

        try:
            wb = load_workbook(str(path), data_only=False)
        except Exception as exc:
            return _failed(method, expected, None, f"Cannot re-open workbook: {exc}")
        try:
            if sheet_name is not None and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.active
            title = ws.title
            observed = _normalize_value(ws[cell_ref].value)
        finally:
            wb.close()

        if _values_match(expected, observed):
            return _passed(method, expected, observed, f"Write verified: {title}!{cell_ref} == {expected!r}")
        return _failed(
            method, expected, observed,
            f"Cell value mismatch at {title}!{cell_ref}: expected {expected!r}, observed {observed!r}",
        )


#: Maps each modifying `spreadsheet.*` type to its verifier class. Read-only
#: types are intentionally absent (→ VerificationRegistry returns SKIPPED).
SPREADSHEET_VERIFIERS: dict[str, type[Verifier]] = {
    "spreadsheet.write_cell": SpreadsheetWriteCellVerifier,
}


def register_spreadsheet_verifiers(registry, *, override: bool = False) -> None:
    """Register a verifier for every modifying `spreadsheet.*` action type."""
    for action_type, verifier_cls in SPREADSHEET_VERIFIERS.items():
        registry.register_verifier(action_type, verifier_cls(), override=override)
