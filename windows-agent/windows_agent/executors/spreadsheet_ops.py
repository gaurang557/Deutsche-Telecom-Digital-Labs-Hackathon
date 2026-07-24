"""`SpreadsheetExecutor` — the `.xlsx` spreadsheet executor (Milestone 4).

WHAT THIS IS
------------
A single executor that performs the `spreadsheet.*` semantic actions against
`.xlsx` workbooks via **openpyxl** (a structured spreadsheet API — far more
reliable than scraping a viewer's GUI; see ARCHITECTURE §3 executor-preference
order). Like `FileExecutor` / `PdfExecutor`, it is deliberately "dumb" about
safety: it does NOT decide permissions or ask for confirmation. That is the
Policy Engine's job (the dispatcher only runs this executor after an ALLOW
decision). See `executors/base.py`.

ACTION VOCABULARY (this milestone)
----------------------------------
Read-only (no side effects, verifier SKIPPED):
  * spreadsheet.list_sheets — the workbook's sheet names
  * spreadsheet.dimensions  — a sheet's used bounds (max row/col + "A1:C10")
  * spreadsheet.read_cell   — one cell's value
  * spreadsheet.read_range  — a rectangular block of values (bounded)

Modifying (verified by re-observation in
`verification/spreadsheet_verifiers.py`):
  * spreadsheet.write_cell  — set one cell's value (creates the workbook if
                              missing; refuses to clobber an occupied cell
                              unless overwrite=true — mirrors file.write_text)

PARAMETER CONVENTIONS
---------------------
`action.target` is the PRIMARY path (the workbook `.xlsx`). Everything else
lives in `action.parameters`, e.g. {"sheet": "Data"}, {"cell": "B7"},
{"range": "A1:C10"}, {"value": 42}, {"overwrite": true}. `sheet` is optional and
defaults to the active/first sheet. Cell/range references are the usual A1-style
strings and are validated (fail closed on a malformed reference).

SHEET RESOLUTION (why it is lenient, and exactly how far)
--------------------------------------------------------
A caller cannot always know a workbook's sheet names before opening it, and a
plan that names a sheet it never observed used to fail the whole task. So
`resolve_sheet_name` resolves the requested name deterministically — no
guessing, no model judgement — in this order: exact match, then match ignoring
surrounding whitespace and letter case, then the workbook's only sheet when it
has exactly one. A name matching nothing in a workbook that has SEVERAL sheets
is still `sheet_not_found`, listing what is available: picking one of several
could silently write to the wrong place. Any resolution that changed the
requested name is reported as evidence (`requested_sheet` / `sheet_substituted`)
so the substitution is auditable instead of invisible.

data_only TRADE-OFF (why reads use data_only=True, writes use data_only=False)
------------------------------------------------------------------------------
openpyxl can load a workbook two ways:
  * data_only=True  → formula cells return the *cached* value Excel last stored.
    If a workbook was written programmatically and never opened/recalculated in
    Excel, that cache is empty and a formula cell reads back as None.
  * data_only=False → formula cells return the *formula string* ("=A1+B1").
We choose **data_only=True for the read actions** because a planner asking to
"read a cell" almost always wants the value, not the formula text; the accepted
trade-off is that an uncalculated formula reads as None (documented in the
ACTION_REFERENCE). **Writes load with data_only=False** so we never destroy
formulas living in *other* cells when we save the workbook back. The write
verifier also re-opens with data_only=False because it re-reads a literal value
we just wrote (no cached-formula ambiguity there).

WHY ASYNC + to_thread
---------------------
The executor contract is async, but openpyxl's load/save calls are blocking
(they parse/serialise a zip of XML on disk). Each operation is dispatched to a
worker thread via `asyncio.to_thread` instead of blocking the event loop —
exactly as file_ops/pdf_ops do for their blocking I/O.

SAFETY / BOUNDING
-----------------
  * `spreadsheet.read_range` is capped at `_RANGE_CELL_CAP` total cells so a
    huge range can never be slurped into memory / evidence; `truncated=true`
    marks a clipped result. The dispatcher additionally bounds evidence.
  * Cell values are normalised to JSON-serialisable primitives (numbers / str /
    bool / None); dates/times become ISO-8601 strings.
  * `write_cell` refuses to overwrite a non-empty cell unless `overwrite=true`
    (mirrors `file.write_text`), so an accidental clobber fails closed.
  * Expected errors are returned as `ExecutorResult(success=False, error=...)` —
    the executor never raises for ordinary failures. (Unexpected exceptions are
    still contained by the dispatcher.) Workbooks are always closed via
    try/finally.
"""

from __future__ import annotations

import asyncio
import datetime as _dt
from collections.abc import Sequence
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, range_boundaries

from ..contracts import Action, ActionError, ErrorCode, ExecutorResult
from .base import BaseExecutor

# Domain-specific error codes. `ActionError.code` is a free string (see
# contracts/error.py); we use stable spreadsheet-specific codes here and fall
# back to the shared ErrorCode values where they fit.
ERR_FILE_NOT_FOUND = "file_not_found"
ERR_NOT_A_SPREADSHEET = "not_a_spreadsheet"
ERR_SHEET_NOT_FOUND = "sheet_not_found"
ERR_INVALID_CELL = "invalid_cell"
ERR_INVALID_RANGE = "invalid_range"
ERR_CELL_OCCUPIED = "cell_occupied"
ERR_INVALID_PARAMS = "invalid_parameters"

#: Hard cap on the number of cells `spreadsheet.read_range` returns, so a huge
#: range can never bloat memory/evidence. `truncated=true` marks a clipped read.
_RANGE_CELL_CAP = 10_000

#: Every action type this executor handles and its deterministic verification
#: requirement.
SPREADSHEET_ACTION_REQUIREMENTS: dict[str, bool] = {
    "spreadsheet.list_sheets": False,
    "spreadsheet.dimensions": False,
    "spreadsheet.read_cell": False,
    "spreadsheet.read_range": False,
    "spreadsheet.write_cell": True,
}
SPREADSHEET_ACTION_TYPES: tuple[str, ...] = tuple(SPREADSHEET_ACTION_REQUIREMENTS)


def _err(code: str, message: str, *, retryable: bool = False) -> ExecutorResult:
    return ExecutorResult(
        success=False,
        error=ActionError(code=code, message=message, retryable=retryable),
    )


def _normalize_value(value: Any) -> Any:
    """Coerce an openpyxl cell value to a JSON-serialisable primitive.

    Numbers / str / bool / None pass through unchanged; datetime/date/time
    become ISO-8601 strings so evidence stays JSON-serialisable.
    """
    if isinstance(value, (_dt.datetime, _dt.date, _dt.time)):
        return value.isoformat()
    return value


def _is_empty(value: Any) -> bool:
    """A cell counts as empty when it holds nothing or an empty string."""
    return value is None or (isinstance(value, str) and value == "")


def resolve_sheet_name(sheet_names: Sequence[str], requested: str) -> str | None:
    """Map a non-empty requested sheet name onto a sheet the workbook has.

    Returns the existing sheet name to use, or None when the request cannot be
    satisfied without guessing. An absent request is the caller's business: it
    means "the active sheet" and never reaches here.

    The order below is fixed and deterministic (see the module's SHEET
    RESOLUTION note). Shared with `verification/spreadsheet_verifiers.py` so a
    write and its independent re-read always agree on which sheet was meant.
    """
    if requested in sheet_names:
        return requested
    wanted = requested.strip().casefold()
    for name in sheet_names:
        if name.strip().casefold() == wanted:
            return name
    if len(sheet_names) == 1:
        return sheet_names[0]
    return None


class SpreadsheetExecutor(BaseExecutor):
    """Executes the `spreadsheet.*` action vocabulary via openpyxl (.xlsx)."""

    name = "spreadsheet"

    async def execute(self, action: Action) -> ExecutorResult:
        # Route on the semantic type. The registry only sends us spreadsheet.*
        # types it was told to, but we still guard against an unexpected one.
        handler = {
            "spreadsheet.list_sheets": self._list_sheets,
            "spreadsheet.dimensions": self._dimensions,
            "spreadsheet.read_cell": self._read_cell,
            "spreadsheet.read_range": self._read_range,
            "spreadsheet.write_cell": self._write_cell,
        }.get(action.type)
        if handler is None:
            return _err(
                ErrorCode.NOT_IMPLEMENTED.value,
                f"SpreadsheetExecutor does not handle action type {action.type!r}",
            )
        # Blocking openpyxl load/save runs off the event loop.
        return await asyncio.to_thread(handler, action)

    # ── helpers ────────────────────────────────────────────────────────────
    @staticmethod
    def _require_target(action: Action) -> Path | None:
        return Path(action.target) if action.target else None

    @staticmethod
    def _param(action: Action, key: str, default: Any = None) -> Any:
        return action.parameters.get(key, default)

    def _open_workbook(
        self, path: Path | None, *, data_only: bool
    ) -> tuple[Any, ExecutorResult | None]:
        """Open + guard an existing workbook, returning (wb, None) or
        (None, error_result).

        Fails closed on a missing path, a non-file target, a non-`.xlsx`
        extension, and anything openpyxl cannot parse as a workbook. The caller
        owns closing the returned workbook.
        """
        if path is None:
            return None, _err(ERR_INVALID_PARAMS, "spreadsheet action requires a target path")
        if not path.exists():
            return None, _err(ERR_FILE_NOT_FOUND, f"File not found: {path}")
        if not path.is_file():
            return None, _err(ERR_NOT_A_SPREADSHEET, f"Not a file: {path}")
        if path.suffix.lower() != ".xlsx":
            return None, _err(ERR_NOT_A_SPREADSHEET, f"Not a .xlsx workbook: {path}")
        try:
            wb = load_workbook(str(path), data_only=data_only)
        except Exception as exc:  # openpyxl raises various zip/XML/format errors
            return None, _err(ERR_NOT_A_SPREADSHEET, f"Cannot open as .xlsx: {path}: {exc}")
        return wb, None

    def _resolve_sheet(self, wb: Any, sheet_name: Any) -> tuple[Any, ExecutorResult | None]:
        """Pick the target worksheet, defaulting to the active/first sheet.

        An absent or blank `sheet` means the active/first sheet. A requested name
        goes through `resolve_sheet_name`; one that cannot be resolved without
        guessing still fails closed with `sheet_not_found`.
        """
        if sheet_name is None:
            return wb.active, None
        requested = sheet_name if isinstance(sheet_name, str) else str(sheet_name)
        if not requested.strip():
            return wb.active, None
        title = resolve_sheet_name(wb.sheetnames, requested)
        if title is None:
            return None, _err(
                ERR_SHEET_NOT_FOUND,
                f"Sheet {sheet_name!r} not found (available: {wb.sheetnames})",
            )
        return wb[title], None

    @staticmethod
    def _sheet_evidence(action: Action, ws: Any) -> dict[str, Any]:
        """Evidence naming the sheet used, and the requested name if it differed.

        A substitution that left no trace would be indistinguishable from the
        caller having named the sheet correctly, so it is reported here for the
        audit trail rather than applied silently.
        """
        evidence: dict[str, Any] = {"sheet": ws.title}
        requested = action.parameters.get("sheet")
        if isinstance(requested, str) and requested.strip() and requested != ws.title:
            evidence["requested_sheet"] = requested
            evidence["sheet_substituted"] = True
        return evidence

    # ── read-only operations ────────────────────────────────────────────────
    def _list_sheets(self, action: Action) -> ExecutorResult:
        path = self._require_target(action)
        wb, err = self._open_workbook(path, data_only=True)
        if err is not None:
            return err
        try:
            return ExecutorResult(
                success=True,
                evidence={"path": str(path), "sheets": list(wb.sheetnames)},
            )
        finally:
            wb.close()

    def _dimensions(self, action: Action) -> ExecutorResult:
        path = self._require_target(action)
        wb, err = self._open_workbook(path, data_only=True)
        if err is not None:
            return err
        try:
            ws, err = self._resolve_sheet(wb, self._param(action, "sheet"))
            if err is not None:
                return err
            max_row = ws.max_row
            max_col = ws.max_column
            dimensions = f"A1:{get_column_letter(max_col)}{max_row}"
            return ExecutorResult(
                success=True,
                evidence={
                    "path": str(path),
                    **self._sheet_evidence(action, ws),
                    "max_row": max_row,
                    "max_col": max_col,
                    "dimensions": dimensions,
                },
            )
        finally:
            wb.close()

    def _read_cell(self, action: Action) -> ExecutorResult:
        path = self._require_target(action)
        cell_ref = self._param(action, "cell")
        if not isinstance(cell_ref, str) or not cell_ref.strip():
            return _err(ERR_INVALID_PARAMS, "spreadsheet.read_cell requires parameters.cell (e.g. 'B7')")
        wb, err = self._open_workbook(path, data_only=True)
        if err is not None:
            return err
        try:
            ws, err = self._resolve_sheet(wb, self._param(action, "sheet"))
            if err is not None:
                return err
            try:
                coordinate_from_string(cell_ref)  # validates the A1 reference
            except Exception:
                return _err(ERR_INVALID_CELL, f"Invalid cell reference: {cell_ref!r}")
            value = _normalize_value(ws[cell_ref].value)
            return ExecutorResult(
                success=True,
                evidence={
                    "path": str(path),
                    **self._sheet_evidence(action, ws),
                    "cell": cell_ref,
                    "value": value,
                },
            )
        finally:
            wb.close()

    def _read_range(self, action: Action) -> ExecutorResult:
        path = self._require_target(action)
        range_ref = self._param(action, "range")
        if not isinstance(range_ref, str) or not range_ref.strip():
            return _err(ERR_INVALID_PARAMS, "spreadsheet.read_range requires parameters.range (e.g. 'A1:C10')")
        wb, err = self._open_workbook(path, data_only=True)
        if err is not None:
            return err
        try:
            ws, err = self._resolve_sheet(wb, self._param(action, "sheet"))
            if err is not None:
                return err
            try:
                min_col, min_row, max_col, max_row = range_boundaries(range_ref)
            except Exception:
                return _err(ERR_INVALID_RANGE, f"Invalid range reference: {range_ref!r}")

            ncols = max_col - min_col + 1
            nrows = max_row - min_row + 1
            # Cap total cells; clip whole rows so the block stays rectangular.
            # (A single row wider than the cap is still returned in full — the
            # cap is a soft guard against unbounded ranges, not a hard cell count.)
            max_rows_allowed = max(1, _RANGE_CELL_CAP // ncols)
            truncated = nrows > max_rows_allowed
            rows_to_read = min(nrows, max_rows_allowed)

            values: list[list[Any]] = []
            for row in ws.iter_rows(
                min_row=min_row,
                max_row=min_row + rows_to_read - 1,
                min_col=min_col,
                max_col=max_col,
                values_only=True,
            ):
                values.append([_normalize_value(v) for v in row])

            return ExecutorResult(
                success=True,
                evidence={
                    "path": str(path),
                    **self._sheet_evidence(action, ws),
                    "range": range_ref,
                    "values": values,
                    "rows": len(values),
                    "cols": ncols,
                    "truncated": truncated,
                },
            )
        finally:
            wb.close()

    # ── modifying operation ─────────────────────────────────────────────────
    def _write_cell(self, action: Action) -> ExecutorResult:
        path = self._require_target(action)
        if path is None:
            return _err(ERR_INVALID_PARAMS, "spreadsheet.write_cell requires a target path")
        if path.suffix.lower() != ".xlsx":
            return _err(ERR_NOT_A_SPREADSHEET, f"Not a .xlsx workbook: {path}")
        cell_ref = self._param(action, "cell")
        if not isinstance(cell_ref, str) or not cell_ref.strip():
            return _err(ERR_INVALID_PARAMS, "spreadsheet.write_cell requires parameters.cell (e.g. 'B7')")
        if "value" not in action.parameters:
            return _err(ERR_INVALID_PARAMS, "spreadsheet.write_cell requires parameters.value")
        value = self._param(action, "value")
        overwrite = bool(self._param(action, "overwrite", False))
        sheet_name = self._param(action, "sheet")

        try:
            coordinate_from_string(cell_ref)  # validates the A1 reference
        except Exception:
            return _err(ERR_INVALID_CELL, f"Invalid cell reference: {cell_ref!r}")

        created = False
        wb: Any = None
        try:
            if not path.exists():
                # New workbook: use its default sheet, optionally renamed to the
                # requested sheet name. We do NOT fail on a missing sheet here —
                # the workbook itself is being created for this write.
                if not path.parent.exists():
                    return _err(ERR_FILE_NOT_FOUND, f"Directory does not exist: {path.parent}")
                wb = Workbook()
                created = True
                ws = wb.active
                if isinstance(sheet_name, str) and sheet_name:
                    ws.title = sheet_name
            else:
                wb, err = self._open_workbook(path, data_only=False)
                if err is not None:
                    return err
                # Existing workbook: a named-but-missing sheet fails closed (do
                # NOT silently create it).
                ws, err = self._resolve_sheet(wb, sheet_name)
                if err is not None:
                    return err

            cell = ws[cell_ref]
            previous = _normalize_value(cell.value)
            occupied = not _is_empty(cell.value)
            if occupied and not overwrite:
                return _err(
                    ERR_CELL_OCCUPIED,
                    f"Cell {cell_ref} on sheet {ws.title!r} is occupied "
                    f"(set overwrite=true to replace): {previous!r}",
                )

            # Keep the value's natural type (do not coerce a JSON number to str).
            cell.value = value
            wb.save(str(path))

            return ExecutorResult(
                success=True,
                evidence={
                    "path": str(path),
                    **self._sheet_evidence(action, ws),
                    "cell": cell_ref,
                    "value": _normalize_value(value),
                    "previous": previous,
                    "created": created,
                    "overwrote": occupied,
                },
                side_effects=[{"type": "spreadsheet.cell_written", "target": str(path)}],
            )
        finally:
            if wb is not None:
                wb.close()


def register_spreadsheet_executor(
    registry, executor: SpreadsheetExecutor | None = None, *, override: bool = False
) -> SpreadsheetExecutor:
    """Register a single `SpreadsheetExecutor` for every `spreadsheet.*` type.

    Returns the executor instance so callers can reuse it. One instance handles
    all spreadsheet operations (it is stateless), matching the "one executor,
    many action types" pattern noted in `executors/base.py`.
    """
    executor = executor or SpreadsheetExecutor()
    for action_type, requires_verification in SPREADSHEET_ACTION_REQUIREMENTS.items():
        registry.register_action(
            action_type,
            executor,
            requires_verification=requires_verification,
            override=override,
        )
    return executor
