"""Milestone 4 — spreadsheet.* executor + write verifier.

Two layers of tests:
  1. The SpreadsheetExecutor in isolation (each action's success + failure paths,
     the overwrite guard, and read-range bounding/truncation).
  2. The full pipeline via the Dispatcher (AllowAllPolicy + spreadsheet verifiers),
     proving execution AND independent re-read verification agree end-to-end.

`.xlsx` fixtures are built on the fly with openpyxl in `tmp_path` (pytest
builtin), so no binary fixtures are committed and the real openpyxl load/save
path is exercised safely.
"""

from pathlib import Path

from openpyxl import Workbook, load_workbook

from windows_agent.contracts import Action, ActionStatus, ExecutorResult, VerificationStatus
from windows_agent.execution import ActionRegistry, Dispatcher
from windows_agent.executors.spreadsheet_ops import SpreadsheetExecutor, register_spreadsheet_executor
from windows_agent.policy import AllowAllPolicy
from windows_agent.verification import (
    SpreadsheetWriteCellVerifier,
    VerificationRegistry,
    register_spreadsheet_verifiers,
)


def _action(type_: str, target=None, parameters=None) -> Action:
    return Action(
        action_id="a1",
        task_id="t1",
        sequence=0,
        type=type_,
        target=str(target) if target is not None else None,
        parameters=parameters or {},
        reason="test",
    )


def _make_workbook(path: Path) -> None:
    """Build a small known workbook: a 'Data' sheet with values + a 'Notes' sheet."""
    wb = Workbook()
    data = wb.active
    data.title = "Data"
    data["A1"] = "Name"
    data["B1"] = "Amount"
    data["A2"] = "Widget"
    data["B2"] = 42
    data["A3"] = "Gadget"
    data["B3"] = 7.5
    notes = wb.create_sheet("Notes")
    notes["A1"] = "hello"
    wb.save(str(path))
    wb.close()


# ── SpreadsheetExecutor unit tests: read-only ──────────────────────────────
async def test_list_sheets(tmp_path):
    wb_path = tmp_path / "book.xlsx"
    _make_workbook(wb_path)
    ex = SpreadsheetExecutor()
    res = await ex.execute(_action("spreadsheet.list_sheets", wb_path))
    assert res.success is True
    assert res.evidence["sheets"] == ["Data", "Notes"]


async def test_dimensions_default_sheet(tmp_path):
    wb_path = tmp_path / "book.xlsx"
    _make_workbook(wb_path)
    ex = SpreadsheetExecutor()
    res = await ex.execute(_action("spreadsheet.dimensions", wb_path))
    assert res.success is True
    assert res.evidence["sheet"] == "Data"
    assert res.evidence["max_row"] == 3
    assert res.evidence["max_col"] == 2
    assert res.evidence["dimensions"] == "A1:B3"


async def test_read_cell_default_and_named_sheet(tmp_path):
    wb_path = tmp_path / "book.xlsx"
    _make_workbook(wb_path)
    ex = SpreadsheetExecutor()
    res = await ex.execute(_action("spreadsheet.read_cell", wb_path, {"cell": "B2"}))
    assert res.success is True
    assert res.evidence["sheet"] == "Data"
    assert res.evidence["value"] == 42
    # A specific (non-active) sheet.
    res2 = await ex.execute(
        _action("spreadsheet.read_cell", wb_path, {"sheet": "Notes", "cell": "A1"})
    )
    assert res2.success is True
    assert res2.evidence["sheet"] == "Notes"
    assert res2.evidence["value"] == "hello"


async def test_read_range(tmp_path):
    wb_path = tmp_path / "book.xlsx"
    _make_workbook(wb_path)
    ex = SpreadsheetExecutor()
    res = await ex.execute(_action("spreadsheet.read_range", wb_path, {"range": "A1:B3"}))
    assert res.success is True
    assert res.evidence["rows"] == 3
    assert res.evidence["cols"] == 2
    assert res.evidence["truncated"] is False
    assert res.evidence["values"][0] == ["Name", "Amount"]
    assert res.evidence["values"][1] == ["Widget", 42]


async def test_read_range_truncation(tmp_path):
    # Build a tall single-column sheet, then read a range that exceeds the cap.
    wb_path = tmp_path / "tall.xlsx"
    wb = Workbook()
    ws = wb.active
    for r in range(1, 21):
        ws.cell(row=r, column=1, value=r)
    wb.save(str(wb_path))
    wb.close()

    ex = SpreadsheetExecutor()
    # Monkeypatch the module cap low so we exercise truncation deterministically.
    import windows_agent.executors.spreadsheet_ops as mod

    original = mod._RANGE_CELL_CAP
    mod._RANGE_CELL_CAP = 5
    try:
        res = await ex.execute(_action("spreadsheet.read_range", wb_path, {"range": "A1:A20"}))
    finally:
        mod._RANGE_CELL_CAP = original
    assert res.success is True
    assert res.evidence["truncated"] is True
    assert res.evidence["rows"] == 5  # cap // cols (5 // 1)
    assert res.evidence["values"][0] == [1]


# ── SpreadsheetExecutor unit tests: write_cell ─────────────────────────────
async def test_write_cell_creates_new_workbook(tmp_path):
    wb_path = tmp_path / "new.xlsx"
    ex = SpreadsheetExecutor()
    res = await ex.execute(
        _action("spreadsheet.write_cell", wb_path, {"sheet": "Sheet1", "cell": "A1", "value": "hi"})
    )
    assert res.success is True
    assert res.evidence["created"] is True
    assert res.evidence["overwrote"] is False
    assert res.evidence["previous"] is None
    # Independently confirm on disk.
    wb = load_workbook(str(wb_path))
    assert wb["Sheet1"]["A1"].value == "hi"
    wb.close()


async def test_write_cell_into_empty_cell(tmp_path):
    wb_path = tmp_path / "book.xlsx"
    _make_workbook(wb_path)
    ex = SpreadsheetExecutor()
    res = await ex.execute(
        _action("spreadsheet.write_cell", wb_path, {"sheet": "Data", "cell": "C1", "value": 100})
    )
    assert res.success is True
    assert res.evidence["created"] is False
    assert res.evidence["overwrote"] is False
    assert res.evidence["value"] == 100
    wb = load_workbook(str(wb_path))
    assert wb["Data"]["C1"].value == 100
    wb.close()


async def test_write_cell_occupied_without_overwrite_fails(tmp_path):
    wb_path = tmp_path / "book.xlsx"
    _make_workbook(wb_path)
    ex = SpreadsheetExecutor()
    res = await ex.execute(
        _action("spreadsheet.write_cell", wb_path, {"sheet": "Data", "cell": "B2", "value": 999})
    )
    assert res.success is False
    assert res.error.code == "cell_occupied"
    # The original value must be untouched.
    wb = load_workbook(str(wb_path))
    assert wb["Data"]["B2"].value == 42
    wb.close()


async def test_write_cell_overwrite_reports_previous(tmp_path):
    wb_path = tmp_path / "book.xlsx"
    _make_workbook(wb_path)
    ex = SpreadsheetExecutor()
    res = await ex.execute(
        _action(
            "spreadsheet.write_cell",
            wb_path,
            {"sheet": "Data", "cell": "B2", "value": 999, "overwrite": True},
        )
    )
    assert res.success is True
    assert res.evidence["overwrote"] is True
    assert res.evidence["previous"] == 42
    assert res.evidence["value"] == 999
    wb = load_workbook(str(wb_path))
    assert wb["Data"]["B2"].value == 999
    wb.close()


async def test_write_cell_keeps_numeric_type(tmp_path):
    wb_path = tmp_path / "num.xlsx"
    ex = SpreadsheetExecutor()
    await ex.execute(_action("spreadsheet.write_cell", wb_path, {"cell": "A1", "value": 3.14}))
    wb = load_workbook(str(wb_path))
    val = wb.active["A1"].value
    wb.close()
    assert isinstance(val, float) and val == 3.14


# ── Error paths ─────────────────────────────────────────────────────────────
async def test_read_missing_file(tmp_path):
    ex = SpreadsheetExecutor()
    res = await ex.execute(_action("spreadsheet.list_sheets", tmp_path / "nope.xlsx"))
    assert res.success is False
    assert res.error.code == "file_not_found"


async def test_read_bad_sheet(tmp_path):
    wb_path = tmp_path / "book.xlsx"
    _make_workbook(wb_path)
    ex = SpreadsheetExecutor()
    res = await ex.execute(
        _action("spreadsheet.read_cell", wb_path, {"sheet": "Ghost", "cell": "A1"})
    )
    assert res.success is False
    assert res.error.code == "sheet_not_found"


async def test_read_bad_cell_ref(tmp_path):
    wb_path = tmp_path / "book.xlsx"
    _make_workbook(wb_path)
    ex = SpreadsheetExecutor()
    res = await ex.execute(_action("spreadsheet.read_cell", wb_path, {"cell": "not-a-cell"}))
    assert res.success is False
    assert res.error.code == "invalid_cell"


async def test_read_bad_range_ref(tmp_path):
    wb_path = tmp_path / "book.xlsx"
    _make_workbook(wb_path)
    ex = SpreadsheetExecutor()
    res = await ex.execute(_action("spreadsheet.read_range", wb_path, {"range": "@@@"}))
    assert res.success is False
    assert res.error.code == "invalid_range"


async def test_non_xlsx_file_fails(tmp_path):
    fake = tmp_path / "data.txt"
    fake.write_text("not a workbook")
    ex = SpreadsheetExecutor()
    res = await ex.execute(_action("spreadsheet.list_sheets", fake))
    assert res.success is False
    assert res.error.code == "not_a_spreadsheet"


async def test_write_missing_value_param(tmp_path):
    wb_path = tmp_path / "book.xlsx"
    _make_workbook(wb_path)
    ex = SpreadsheetExecutor()
    res = await ex.execute(_action("spreadsheet.write_cell", wb_path, {"cell": "C1"}))
    assert res.success is False
    assert res.error.code == "invalid_parameters"


async def test_write_existing_workbook_bad_sheet_fails_closed(tmp_path):
    wb_path = tmp_path / "book.xlsx"
    _make_workbook(wb_path)
    ex = SpreadsheetExecutor()
    res = await ex.execute(
        _action("spreadsheet.write_cell", wb_path, {"sheet": "Ghost", "cell": "A1", "value": 1})
    )
    assert res.success is False
    assert res.error.code == "sheet_not_found"


# ── Verifier unit tests ────────────────────────────────────────────────────
async def test_write_verifier_passes(tmp_path):
    wb_path = tmp_path / "book.xlsx"
    _make_workbook(wb_path)
    ex = SpreadsheetExecutor()
    action = _action("spreadsheet.write_cell", wb_path, {"sheet": "Data", "cell": "C1", "value": 55})
    res = await ex.execute(action)
    assert res.success is True
    vr = await SpreadsheetWriteCellVerifier().verify(action, res)
    assert vr.status == VerificationStatus.PASSED
    assert vr.observed == 55


async def test_write_verifier_passes_int_vs_float(tmp_path):
    # Intended int 10 re-read as 10.0 must still PASS (numeric comparison).
    wb_path = tmp_path / "book.xlsx"
    _make_workbook(wb_path)
    ex = SpreadsheetExecutor()
    action = _action("spreadsheet.write_cell", wb_path, {"cell": "C1", "value": 10})
    res = await ex.execute(action)
    # Feed evidence a float to simulate a float re-read comparison.
    tampered = ExecutorResult(success=True, evidence={**res.evidence, "value": 10.0})
    vr = await SpreadsheetWriteCellVerifier().verify(action, tampered)
    assert vr.status == VerificationStatus.PASSED


async def test_write_verifier_fails_on_disk_change(tmp_path):
    wb_path = tmp_path / "book.xlsx"
    _make_workbook(wb_path)
    ex = SpreadsheetExecutor()
    action = _action("spreadsheet.write_cell", wb_path, {"sheet": "Data", "cell": "C1", "value": 55})
    res = await ex.execute(action)
    # Externally change the cell so re-observation disagrees with the intent.
    wb = load_workbook(str(wb_path))
    wb["Data"]["C1"] = 999
    wb.save(str(wb_path))
    wb.close()
    vr = await SpreadsheetWriteCellVerifier().verify(action, res)
    assert vr.status == VerificationStatus.FAILED
    assert vr.expected == 55
    assert vr.observed == 999


# ── End-to-end via the Dispatcher (policy + execution + verification) ───────
def _pipeline():
    reg = ActionRegistry()
    register_spreadsheet_executor(reg)
    vreg = VerificationRegistry()
    register_spreadsheet_verifiers(vreg)
    return Dispatcher(reg, AllowAllPolicy(), verification=vreg)


async def test_pipeline_write_cell_success_and_verified(tmp_path):
    wb_path = tmp_path / "book.xlsx"
    _make_workbook(wb_path)
    disp = _pipeline()
    result = await disp.dispatch(
        _action("spreadsheet.write_cell", wb_path, {"sheet": "Data", "cell": "C1", "value": "done"})
    )
    assert result.status == ActionStatus.SUCCESS
    assert result.verification.status == VerificationStatus.PASSED
    wb = load_workbook(str(wb_path))
    assert wb["Data"]["C1"].value == "done"
    wb.close()


async def test_pipeline_read_cell_skips_verification(tmp_path):
    wb_path = tmp_path / "book.xlsx"
    _make_workbook(wb_path)
    disp = _pipeline()
    result = await disp.dispatch(_action("spreadsheet.read_cell", wb_path, {"cell": "B2"}))
    assert result.status == ActionStatus.SUCCESS
    # Read-only → no verifier registered → SKIPPED.
    assert result.verification.status == VerificationStatus.SKIPPED


async def test_register_spreadsheet_executor_covers_all_types():
    reg = ActionRegistry()
    register_spreadsheet_executor(reg)
    for t in (
        "spreadsheet.list_sheets",
        "spreadsheet.dimensions",
        "spreadsheet.read_cell",
        "spreadsheet.read_range",
        "spreadsheet.write_cell",
    ):
        assert reg.has_action(t)
