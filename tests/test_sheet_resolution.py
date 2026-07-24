"""Sheet resolution — a plan may not know a workbook's sheet names.

A live run of the PDF→workbook workflow failed at its third step: the planner had
emitted `spreadsheet.list_sheets`, ignored the result, and then invented a sheet
name for `spreadsheet.read_range`. The whole task died on
`Sheet 'X' not found (available: ['Y'])`.

The fix is deterministic and lives in the executor, not in the model: a requested
sheet name is matched exactly, then up to whitespace and letter case, then — when
the workbook has exactly ONE sheet — onto that sheet, because nothing else could
have been meant. A name that matches nothing in a workbook with SEVERAL sheets
still fails closed, since choosing among several would risk silently writing to
the wrong place.

Every fixture here is built in `tmp_path` with openpyxl. None of the sheet names,
labels, or file names below is one the production code knows about; the point of
the tests is that the code never needs to know them.
"""

# ruff: noqa: I001

from pathlib import Path
from typing import Any
from uuid import uuid4

import pytest
from openpyxl import Workbook, load_workbook

from app.execution.hybrid import HybridExecutor, _resolve_references, build_structured_dispatcher
from app.planning.normalizer import build_action_plan
from app.schemas import ActionResult, ActionStatus, DraftAction, DraftPlan, TaskRequest

from windows_agent.audit import InMemoryAuditSink
from windows_agent.contracts import Action as StructuredAction
from windows_agent.executors.spreadsheet_ops import SpreadsheetExecutor, resolve_sheet_name


def _structured_action(type_: str, target: Path, parameters: dict[str, Any]) -> StructuredAction:
    return StructuredAction(
        action_id="a1",
        task_id="t1",
        sequence=0,
        type=type_,
        target=str(target),
        parameters=parameters,
        reason="test",
    )


def _single_sheet_workbook(path: Path, title: str) -> None:
    """A workbook whose only sheet is named `title`, with one labelled row."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title
    sheet.append(["Territory", "Takings"])
    sheet.append(["Coastal", None])
    workbook.save(path)
    workbook.close()


def _multi_sheet_workbook(path: Path, titles: tuple[str, ...]) -> None:
    workbook = Workbook()
    first = workbook.active
    first.title = titles[0]
    first["A1"] = titles[0]
    for title in titles[1:]:
        workbook.create_sheet(title)["A1"] = title
    workbook.save(path)
    workbook.close()


def _active_title(path: Path) -> str:
    workbook = load_workbook(path)
    try:
        return workbook.active.title
    finally:
        workbook.close()


# ── the deterministic rule, on its own ────────────────────────────────────────
@pytest.mark.parametrize(
    ("names", "requested", "expected"),
    [
        (["Ledger"], "Ledger", "Ledger"),
        (["Ledger", "Notes"], "Notes", "Notes"),
        (["Ledger", "Notes"], "  ledger ", "Ledger"),
        (["Ledger", "Notes"], "LEDGER", "Ledger"),
        (["Ledger"], "Whatever The Model Said", "Ledger"),
        (["Ledger", "Notes"], "Whatever The Model Said", None),
        (["Ledger", "Notes", "Archive"], "Ledgers", None),
    ],
    ids=[
        "exact",
        "exact_non_active",
        "whitespace_and_case",
        "case_only",
        "only_sheet_is_the_only_candidate",
        "two_sheets_no_match_is_undecidable",
        "three_sheets_near_miss_is_undecidable",
    ],
)
def test_resolve_sheet_name_resolves_or_refuses_but_never_guesses(
    names: list[str],
    requested: str,
    expected: str | None,
) -> None:
    assert resolve_sheet_name(names, requested) == expected


# ── the executor: reads ───────────────────────────────────────────────────────
async def test_omitted_sheet_reads_the_workbooks_active_sheet(tmp_path: Path) -> None:
    """A plan that names no sheet is the shape we now ask the planner for."""
    path = tmp_path / "book.xlsx"
    _multi_sheet_workbook(path, ("Ledger", "Notes"))

    result = await SpreadsheetExecutor().execute(
        _structured_action("spreadsheet.read_range", path, {"range": "A1:B2"})
    )

    assert result.success is True
    assert result.evidence["sheet"] == _active_title(path)
    assert "sheet_substituted" not in result.evidence


async def test_exact_sheet_name_still_selects_that_sheet(tmp_path: Path) -> None:
    path = tmp_path / "book.xlsx"
    _multi_sheet_workbook(path, ("Ledger", "Notes"))

    result = await SpreadsheetExecutor().execute(
        _structured_action("spreadsheet.read_cell", path, {"sheet": "Notes", "cell": "A1"})
    )

    assert result.success is True
    assert result.evidence["sheet"] == "Notes"
    assert result.evidence["value"] == "Notes"
    assert "sheet_substituted" not in result.evidence


async def test_sheet_name_differing_in_case_and_whitespace_resolves(tmp_path: Path) -> None:
    """Two sheets, so only the case-insensitive match can be doing the work."""
    path = tmp_path / "book.xlsx"
    _multi_sheet_workbook(path, ("Ledger", "Notes"))

    result = await SpreadsheetExecutor().execute(
        _structured_action("spreadsheet.read_cell", path, {"sheet": "  notes ", "cell": "A1"})
    )

    assert result.success is True
    assert result.evidence["sheet"] == "Notes"
    assert result.evidence["requested_sheet"] == "  notes "
    assert result.evidence["sheet_substituted"] is True


async def test_invented_sheet_name_in_a_single_sheet_workbook_resolves_and_is_recorded(
    tmp_path: Path,
) -> None:
    """The exact live failure: a name the workbook does not have, one sheet."""
    path = tmp_path / "book.xlsx"
    _single_sheet_workbook(path, "Ledger")

    result = await SpreadsheetExecutor().execute(
        _structured_action(
            "spreadsheet.read_range",
            path,
            {"sheet": "A Name Nobody Observed", "range": "A1:B2"},
        )
    )

    assert result.success is True
    assert result.evidence["sheet"] == "Ledger"
    assert result.evidence["requested_sheet"] == "A Name Nobody Observed"
    assert result.evidence["sheet_substituted"] is True
    assert result.evidence["values"][1] == ["Coastal", None]


async def test_invented_sheet_name_in_a_multi_sheet_workbook_still_fails_closed(
    tmp_path: Path,
) -> None:
    path = tmp_path / "book.xlsx"
    _multi_sheet_workbook(path, ("Ledger", "Notes", "Archive"))

    result = await SpreadsheetExecutor().execute(
        _structured_action(
            "spreadsheet.read_range",
            path,
            {"sheet": "A Name Nobody Observed", "range": "A1:B2"},
        )
    )

    assert result.success is False
    assert result.error is not None
    assert result.error.code == "sheet_not_found"
    assert "A Name Nobody Observed" in result.error.message
    assert "'Ledger', 'Notes', 'Archive'" in result.error.message


# ── the executor: writes, where a wrong sheet would corrupt the workbook ──────
async def test_invented_sheet_name_writes_to_the_only_sheet(tmp_path: Path) -> None:
    path = tmp_path / "book.xlsx"
    _single_sheet_workbook(path, "Ledger")

    result = await SpreadsheetExecutor().execute(
        _structured_action(
            "spreadsheet.write_cell",
            path,
            {"sheet": "A Name Nobody Observed", "cell": "B2", "value": 27.4},
        )
    )

    assert result.success is True
    assert result.evidence["sheet"] == "Ledger"
    assert result.evidence["sheet_substituted"] is True
    workbook = load_workbook(path)
    try:
        assert workbook.sheetnames == ["Ledger"]
        assert workbook["Ledger"]["B2"].value == 27.4
    finally:
        workbook.close()


async def test_invented_sheet_name_never_writes_into_a_multi_sheet_workbook(
    tmp_path: Path,
) -> None:
    """Guessing among several sheets could silently write to the wrong place."""
    path = tmp_path / "book.xlsx"
    _multi_sheet_workbook(path, ("Ledger", "Notes"))

    result = await SpreadsheetExecutor().execute(
        _structured_action(
            "spreadsheet.write_cell",
            path,
            {"sheet": "A Name Nobody Observed", "cell": "B2", "value": 27.4},
        )
    )

    assert result.success is False
    assert result.error is not None
    assert result.error.code == "sheet_not_found"
    workbook = load_workbook(path)
    try:
        assert workbook.sheetnames == ["Ledger", "Notes"]
        assert all(sheet["B2"].value is None for sheet in workbook.worksheets)
    finally:
        workbook.close()


# ── end to end, the failure that started this ─────────────────────────────────
def _invented_sheet_draft(workbook_path: Path, invented: str) -> DraftPlan:
    """A plan shaped exactly like the live failure: a sheet name out of thin air."""
    return DraftPlan(
        summary="I'll look at the workbook and record the figure.",
        actions=[
            DraftAction(
                step_key="layout",
                type="spreadsheet.read_range",
                target=str(workbook_path),
                parameters={"sheet": invented, "range": "A1:B2"},
                description="Look at the rows.",
                expected_result={"contains": "the label"},
            ),
            DraftAction(
                step_key="write",
                type="spreadsheet.write_cell",
                target=str(workbook_path),
                parameters={"sheet": invented, "cell": "B2", "value": 27.4},
                depends_on=["layout"],
                description="Record the figure.",
                expected_result={"written": True},
            ),
        ],
    )


async def test_plan_with_an_invented_sheet_name_completes_and_verifies(
    tmp_path: Path,
) -> None:
    path = tmp_path / "book.xlsx"
    _single_sheet_workbook(path, "Ledger")
    plan = build_action_plan(
        TaskRequest(text=f"Put the coastal figure into {path}"),
        _invented_sheet_draft(path, "A Name Nobody Observed"),
    )
    audit = InMemoryAuditSink()
    executor = HybridExecutor(dispatcher=build_structured_dispatcher(audit), audit=audit)

    response = await executor.execute_plan(plan, set())

    assert response.status == "completed"
    assert [result.status for result in response.results] == ["succeeded", "succeeded"]
    assert response.results[-1].verification is not None
    assert response.results[-1].verification.passed is True
    workbook = load_workbook(path)
    try:
        assert workbook["Ledger"]["B2"].value == 27.4
    finally:
        workbook.close()
    # The leniency is visible rather than silent.
    assert "plan_revised" in [event.event_type.value for event in audit.events]
    substitution = next(
        event for event in audit.events if event.event_type.value == "plan_revised"
    )
    assert "A Name Nobody Observed" in substitution.summary
    assert "Ledger" in substitution.summary


async def test_read_only_step_says_there_was_nothing_to_verify(tmp_path: Path) -> None:
    """A read registers no verifier by design; the wording must not read as a gap."""
    path = tmp_path / "book.xlsx"
    _single_sheet_workbook(path, "Ledger")
    plan = build_action_plan(
        TaskRequest(text=f"Put the coastal figure into {path}"),
        _invented_sheet_draft(path, "Ledger"),
    )

    response = await HybridExecutor().execute_plan(plan, set())

    read_step = response.results[0]
    assert read_step.status == "succeeded"
    assert read_step.verification is not None
    assert read_step.verification.passed is None
    assert "No verifier registered" not in read_step.verification.message
    assert read_step.verification.message == (
        "Nothing to verify: this step only read, it changed nothing."
    )
    # The implementer-facing detail is kept, just not shown as the headline.
    assert read_step.verification.evidence["method"] == "none"


# ── the $ref mechanism was capable of this all along ─────────────────────────
def test_a_list_sheets_result_can_bind_into_a_sheet_parameter() -> None:
    """Documents why the root cause was the plan, not a gap in result binding.

    `spreadsheet.list_sheets` returns a list, and a `$ref` path can index it, so
    the planner *could* have bound step 3's sheet to step 2's result. It emitted a
    literal string instead — which is why the durable fix is in the executor and
    the guidance now tells the planner to omit a sheet it has not observed.
    """
    prior = {
        "sheets": ActionResult(
            action_id=uuid4(),
            status=ActionStatus.SUCCEEDED,
            evidence={"path": "book.xlsx", "sheets": ["Ledger", "Notes"]},
        )
    }

    resolved = _resolve_references({"sheet": {"$ref": "sheets.evidence.sheets.0"}}, prior)

    assert resolved == {"sheet": "Ledger"}
