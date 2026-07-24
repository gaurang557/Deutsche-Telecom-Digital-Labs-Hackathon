"""Path resolution — a spoken request routinely omits the folder a file sits in.

A live run of the PDF→workbook workflow died on its FIRST step. The user said
"in desktop <folder> find the revenue in <report>.pdf"; the planner emitted
`Desktop/<report>.pdf`, dropping the folder the user had named, and the executor
correctly reported that no such file existed.

The durable fix is deterministic and lives in the resolution layer rather than in
the model: when a READ step names a file that is not there, the file's own name is
looked for in a small, bounded region below the root the path already resolved
under, and a single unambiguous match is used and reported. Several matches fail
closed with the candidates named, because choosing one would risk reading the
wrong file. Nothing found leaves the executor's own "File not found" as the
failure, exactly as before.

Two boundaries are load-bearing and asserted here:
  * the walk never leaves the root the request already resolved under, never
    looks above it, and never follows a link or a hidden/system tree out of it;
  * a write or a create destination is NEVER redirected. `spreadsheet.write_cell`
    creating a workbook at the path the plan named is the intended behaviour, and
    a same-named workbook elsewhere must not steal that write.

Every fixture below is built inside `tmp_path`, with a throwaway home directory
patched over `Path.home`, so no test reads or writes a real Desktop. None of the
folder or file names used here is one the production code knows about — that is
the point: the code never needs to know them.
"""

# ruff: noqa: I001

import os
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from app.execution import hybrid
from app.execution.hybrid import (
    HybridExecutor,
    _containing_root,
    _discover_readable_file,
    _resolve_local_path,
    _resolve_path_parameters,
    build_structured_dispatcher,
)
from app.planning.exceptions import InvalidPlannerResponseError
from app.planning.normalizer import build_action_plan
from app.schemas import DraftAction, DraftPlan, TaskRequest

from windows_agent.audit import InMemoryAuditSink


@pytest.fixture
def desktop(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> Path:
    """A throwaway home whose Desktop is the only real allowlisted root in play."""
    home = tmp_path / "home"
    (home / "Desktop").mkdir(parents=True)
    monkeypatch.setattr(Path, "home", classmethod(lambda cls: home))
    return home / "Desktop"


def _same(left: str | Path, right: str | Path) -> bool:
    """Path equality the way the OS sees it, so casing never decides a test."""
    return os.path.normcase(str(left)) == os.path.normcase(str(right))


def _workbook(path: Path, label: str) -> None:
    """A one-sheet workbook with a labelled row and an empty value cell."""
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Territory", "Takings"])
    sheet.append([label, None])
    workbook.save(path)
    workbook.close()


def _workbook_bytes(path: Path, label: str) -> bytes:
    """A workbook, plus its exact bytes, so a test can prove it was untouched."""
    _workbook(path, label)
    return path.read_bytes()


def _touch(path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(b"not really a workbook")
    return path


def _single_step_plan(type_: str, target: str, parameters: dict) -> DraftPlan:
    return DraftPlan(
        summary="I'll take care of that figure.",
        actions=[
            DraftAction(
                step_key="step",
                type=type_,
                target=target,
                parameters=parameters,
                description="Handle the figure.",
                expected_result={"contains": "the label"},
            )
        ],
    )


# ── an exact path is still an exact path ──────────────────────────────────────
def test_a_path_that_exists_is_left_exactly_as_written(desktop: Path) -> None:
    _touch(desktop / "ledger.xlsx")

    assert _discover_readable_file(_resolve_local_path("Desktop/ledger.xlsx")) is None


def test_a_nested_path_the_plan_got_right_is_left_alone(desktop: Path) -> None:
    """The whole point of the prompt rule: a correct folder segment needs no help."""
    actual = _touch(desktop / "quarter three" / "ledger.xlsx")

    resolved = _resolve_local_path("Desktop/quarter three/ledger.xlsx")

    assert _same(resolved, actual)
    assert _discover_readable_file(resolved) is None


# ── the dropped folder segment, recovered ─────────────────────────────────────
def test_a_file_one_folder_deeper_is_found(desktop: Path) -> None:
    actual = _touch(desktop / "handover" / "ledger.xlsx")

    found = _discover_readable_file(_resolve_local_path("Desktop/ledger.xlsx"))

    assert found is not None
    assert _same(found, actual)


def test_the_name_is_matched_regardless_of_letter_case(desktop: Path) -> None:
    actual = _touch(desktop / "handover" / "ledger.xlsx")

    found = _discover_readable_file(_resolve_local_path("Desktop/LEDGER.XLSX"))

    assert found is not None
    assert _same(found, actual)


def test_a_file_that_is_nowhere_below_the_root_is_left_to_fail(desktop: Path) -> None:
    """Zero matches must not change the request, so the reported error is unchanged."""
    assert _discover_readable_file(_resolve_local_path("Desktop/ledger.xlsx")) is None


# ── several candidates fail closed ────────────────────────────────────────────
def test_two_files_with_the_same_name_fail_with_both_named(desktop: Path) -> None:
    _touch(desktop / "one" / "ledger.xlsx")
    _touch(desktop / "one" / "two" / "ledger.xlsx")

    with pytest.raises(ValueError) as excinfo:
        _discover_readable_file(_resolve_local_path("Desktop/ledger.xlsx"))

    message = str(excinfo.value)
    assert "Desktop/one/ledger.xlsx" in message
    assert "Desktop/one/two/ledger.xlsx" in message
    assert "which folder you meant" in message


def test_a_candidate_list_is_bounded(desktop: Path) -> None:
    for index in range(hybrid._DISCOVERY_MAX_CANDIDATES + 4):
        _touch(desktop / f"copy{index:02d}" / "ledger.xlsx")

    with pytest.raises(ValueError) as excinfo:
        _discover_readable_file(_resolve_local_path("Desktop/ledger.xlsx"))

    # Each candidate is named once, under its own uniquely named folder.
    assert str(excinfo.value).count("copy") == hybrid._DISCOVERY_MAX_CANDIDATES


# ── the search cannot leave the root it was given ─────────────────────────────
def test_a_root_is_never_itself_a_thing_to_search_inside_of(desktop: Path) -> None:
    assert _containing_root(desktop) is None
    assert _same(_containing_root(desktop / "gone.xlsx"), desktop)


def test_a_parent_traversal_lands_outside_every_root_and_is_not_searched(
    desktop: Path,
) -> None:
    """`..` collapses during resolution, so the request leaves the allowlist."""
    _touch(desktop.parent / "private" / "ledger.xlsx")

    resolved = _resolve_local_path("Desktop/../private/ledger.xlsx")

    assert not Path(resolved).is_relative_to(desktop)
    assert _containing_root(Path(resolved)) is None
    assert _discover_readable_file(resolved) is None


def test_the_search_never_looks_above_the_root(desktop: Path) -> None:
    """A same-named file in the home directory itself must stay invisible."""
    _touch(desktop.parent / "ledger.xlsx")

    assert _discover_readable_file(_resolve_local_path("Desktop/ledger.xlsx")) is None


def test_a_path_outside_every_root_is_never_searched(
    desktop: Path, tmp_path: Path
) -> None:
    elsewhere = _touch(tmp_path / "elsewhere" / "ledger.xlsx").parent

    assert _discover_readable_file(str(elsewhere / "missing" / "ledger.xlsx")) is None


def test_a_hidden_folder_is_not_walked(desktop: Path) -> None:
    _touch(desktop / ".cache" / "ledger.xlsx")

    assert _discover_readable_file(_resolve_local_path("Desktop/ledger.xlsx")) is None


def test_a_link_pointing_out_of_the_root_is_not_followed(
    desktop: Path, tmp_path: Path
) -> None:
    outside = _touch(tmp_path / "outside" / "ledger.xlsx").parent
    try:
        (desktop / "shortcut").symlink_to(outside, target_is_directory=True)
    except (OSError, NotImplementedError):
        pytest.skip("creating a directory symlink needs privileges this run lacks")

    assert _discover_readable_file(_resolve_local_path("Desktop/ledger.xlsx")) is None


# ── the bounds ────────────────────────────────────────────────────────────────
def test_the_bounds_stay_small() -> None:
    """A regression here would turn a bounded lookup into a filesystem crawl."""
    assert hybrid._DISCOVERY_MAX_DEPTH <= 3
    assert hybrid._DISCOVERY_MAX_DIRECTORIES <= 500
    assert hybrid._DISCOVERY_MAX_CANDIDATES <= 20


def test_a_file_at_the_depth_bound_is_found(desktop: Path) -> None:
    parts = ["level"] * hybrid._DISCOVERY_MAX_DEPTH
    actual = _touch(desktop.joinpath(*parts) / "ledger.xlsx")

    found = _discover_readable_file(_resolve_local_path("Desktop/ledger.xlsx"))

    assert found is not None
    assert _same(found, actual)


def test_a_file_past_the_depth_bound_is_not_found(desktop: Path) -> None:
    parts = ["level"] * (hybrid._DISCOVERY_MAX_DEPTH + 1)
    _touch(desktop.joinpath(*parts) / "ledger.xlsx")

    assert _discover_readable_file(_resolve_local_path("Desktop/ledger.xlsx")) is None


def test_the_search_gives_up_when_its_directory_budget_runs_out(
    desktop: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    for name in ("a", "b", "c"):
        (desktop / name).mkdir()
    actual = _touch(desktop / "d" / "ledger.xlsx")

    monkeypatch.setattr(hybrid, "_DISCOVERY_MAX_DIRECTORIES", 2)
    assert _discover_readable_file(_resolve_local_path("Desktop/ledger.xlsx")) is None

    # The same tree resolves once the walk is allowed to reach that far, so the
    # budget is what stopped it and not something about the fixture.
    monkeypatch.setattr(hybrid, "_DISCOVERY_MAX_DIRECTORIES", 250)
    found = _discover_readable_file(_resolve_local_path("Desktop/ledger.xlsx"))
    assert found is not None
    assert _same(found, actual)


# ── writes and destinations are never redirected ──────────────────────────────
def test_a_destination_parameter_is_resolved_but_never_searched(desktop: Path) -> None:
    _touch(desktop / "handover" / "notes.txt")

    resolved = _resolve_path_parameters({"destination": "Desktop/notes.txt"})

    assert _same(resolved["destination"], desktop / "notes.txt")
    assert "handover" not in resolved["destination"]


async def test_a_write_creates_the_path_the_plan_named(desktop: Path) -> None:
    """A create must not be captured by a same-named workbook found elsewhere.

    The plan here only writes. Nothing in it reads that workbook, so there is no
    evidence the file was expected to exist and "create it where I said" is the
    only reading of the request — even though a same-named workbook is sitting one
    folder away. Contrast the next test, where the plan reads the workbook first.
    """
    elsewhere = desktop / "handover" / "ledger.xlsx"
    _workbook(elsewhere, "Coastal")
    plan = build_action_plan(
        TaskRequest(text="record the takings in Desktop/ledger.xlsx"),
        _single_step_plan(
            "spreadsheet.write_cell",
            "Desktop/ledger.xlsx",
            {"cell": "B2", "value": 27.4},
        ),
    )

    response = await HybridExecutor().execute_plan(plan, set())

    assert response.status == "completed"
    assert (desktop / "ledger.xlsx").is_file()
    assert response.results[0].evidence.get("path_substituted") is None
    workbook = load_workbook(elsewhere)
    try:
        assert workbook.active["B2"].value is None
    finally:
        workbook.close()


async def test_a_write_follows_the_workbook_the_same_plan_reads(desktop: Path) -> None:
    """A plan that reads a workbook and then writes to it must mean one file.

    This is the primary workflow's shape: inspect the sheet, then fill a cell in
    it. Reading a file the plan is about to create would be meaningless, so the
    read is proof the workbook is expected to exist — and letting the write create
    a second, empty workbook beside it would report success while leaving the
    user's real file untouched.
    """
    actual = desktop / "handover" / "ledger.xlsx"
    _workbook(actual, "Coastal")
    plan = build_action_plan(
        TaskRequest(text="read Desktop/ledger.xlsx and record the takings in it"),
        DraftPlan(
            summary="I'll look at that sheet and record the takings.",
            actions=[
                DraftAction(
                    step_key="look",
                    type="spreadsheet.read_range",
                    target="Desktop/ledger.xlsx",
                    parameters={"range": "A1:B2"},
                    description="Look at the sheet.",
                    expected_result={"contains": "the label"},
                ),
                DraftAction(
                    step_key="record",
                    type="spreadsheet.write_cell",
                    target="Desktop/ledger.xlsx",
                    parameters={"cell": "B2", "value": 27.4},
                    depends_on=["look"],
                    description="Record the takings.",
                    expected_result={"written": True},
                ),
            ],
        ),
    )

    write = plan.actions[1]
    # Both steps agree on one file, and the write is bound to the resolved one.
    assert _same(write.target, actual)
    assert _same(plan.actions[0].target, actual)
    assert _same(write.resolved_from, desktop / "ledger.xlsx")

    response = await HybridExecutor().execute_plan(plan, set())

    assert response.status == "completed"
    # No phantom workbook was created beside the real one.
    assert not (desktop / "ledger.xlsx").exists()
    workbook = load_workbook(actual)
    try:
        assert workbook.active["B2"].value == 27.4
    finally:
        workbook.close()


def test_a_write_keeps_the_named_path_when_nothing_with_that_name_exists(
    desktop: Path,
) -> None:
    """Zero matches is the genuine create case, even with a read step present.

    Nothing anywhere is called ledger.xlsx, so there is nothing to follow and the
    write must land exactly where the plan said.
    """
    plan = build_action_plan(
        TaskRequest(text="read Desktop/ledger.xlsx and record the takings in it"),
        DraftPlan(
            summary="I'll look at that sheet and record the takings.",
            actions=[
                DraftAction(
                    step_key="look",
                    type="spreadsheet.read_range",
                    target="Desktop/ledger.xlsx",
                    parameters={"range": "A1:B2"},
                    description="Look at the sheet.",
                    expected_result={"contains": "the label"},
                ),
                DraftAction(
                    step_key="record",
                    type="spreadsheet.write_cell",
                    target="Desktop/ledger.xlsx",
                    parameters={"cell": "B2", "value": 27.4},
                    depends_on=["look"],
                    description="Record the takings.",
                    expected_result={"written": True},
                ),
            ],
        ),
    )

    assert plan.actions[1].target == "Desktop/ledger.xlsx"
    assert plan.actions[1].resolved_from is None


def test_an_ambiguous_write_that_the_plan_reads_fails_loudly(desktop: Path) -> None:
    first = _workbook_bytes(desktop / "one" / "ledger.xlsx", "Coastal")
    second = _workbook_bytes(desktop / "two" / "ledger.xlsx", "Inland")

    with pytest.raises(InvalidPlannerResponseError) as excinfo:
        build_action_plan(
            TaskRequest(text="read Desktop/ledger.xlsx and record the takings in it"),
            DraftPlan(
                summary="I'll record the takings.",
                actions=[
                    DraftAction(
                        step_key="look",
                        type="spreadsheet.read_range",
                        target="Desktop/ledger.xlsx",
                        parameters={"range": "A1:B2"},
                        description="Look at the sheet.",
                        expected_result={"contains": "the label"},
                    ),
                    DraftAction(
                        step_key="record",
                        type="spreadsheet.write_cell",
                        target="Desktop/ledger.xlsx",
                        parameters={"cell": "B2", "value": 27.4},
                        depends_on=["look"],
                        description="Record the takings.",
                        expected_result={"written": True},
                    ),
                ],
            ),
        )

    error = str(excinfo.value)
    assert "Desktop/one/ledger.xlsx" in error
    assert "Desktop/two/ledger.xlsx" in error
    assert (desktop / "one" / "ledger.xlsx").read_bytes() == first
    assert (desktop / "two" / "ledger.xlsx").read_bytes() == second
    assert not (desktop / "ledger.xlsx").exists()


# ── end to end, the failure that started this ─────────────────────────────────
async def test_a_read_whose_folder_segment_was_dropped_succeeds_and_says_so(
    desktop: Path,
) -> None:
    actual = desktop / "handover" / "ledger.xlsx"
    _workbook(actual, "Coastal")
    plan = build_action_plan(
        TaskRequest(text="read the takings from Desktop/ledger.xlsx"),
        _single_step_plan(
            "spreadsheet.read_range", "Desktop/ledger.xlsx", {"range": "A1:B2"}
        ),
    )
    audit = InMemoryAuditSink()
    executor = HybridExecutor(dispatcher=build_structured_dispatcher(audit), audit=audit)

    response = await executor.execute_plan(plan, set())

    assert response.status == "completed"
    result = response.results[0]
    assert result.status == "succeeded"
    assert _same(result.evidence["path"], actual)
    # The leniency is visible rather than silent.
    assert result.evidence["path_substituted"] is True
    assert _same(result.evidence["requested_path"], desktop / "ledger.xlsx")
    revisions = [
        event for event in audit.events if event.event_type.value == "plan_revised"
    ]
    assert len(revisions) == 1
    assert revisions[0].outcome == "path_substituted"
    assert "handover" in revisions[0].summary


async def test_a_read_with_nothing_to_find_reports_the_path_it_was_given(
    desktop: Path,
) -> None:
    plan = build_action_plan(
        TaskRequest(text="read the takings from Desktop/ledger.xlsx"),
        _single_step_plan(
            "spreadsheet.read_range", "Desktop/ledger.xlsx", {"range": "A1:B2"}
        ),
    )

    response = await HybridExecutor().execute_plan(plan, set())

    assert response.status == "failed"
    error = response.results[0].error or ""
    assert error.startswith("File not found:")
    assert "ledger.xlsx" in error


def test_an_ambiguous_read_fails_at_plan_time_without_choosing_a_file(
    desktop: Path,
) -> None:
    # Detected while the plan is being built, deliberately: only the user can say
    # which folder they meant, so nothing should reach execution at all.
    first = _workbook_bytes(desktop / "one" / "ledger.xlsx", "Coastal")
    second = _workbook_bytes(desktop / "two" / "ledger.xlsx", "Inland")

    with pytest.raises(InvalidPlannerResponseError) as excinfo:
        build_action_plan(
            TaskRequest(text="read the takings from Desktop/ledger.xlsx"),
            _single_step_plan(
                "spreadsheet.read_range", "Desktop/ledger.xlsx", {"range": "A1:B2"}
            ),
        )

    error = str(excinfo.value)
    # Both candidates must stay named, so a future change cannot quietly drop to
    # picking one of them.
    assert "Desktop/one/ledger.xlsx" in error
    assert "Desktop/two/ledger.xlsx" in error
    # No candidate was opened, read, or altered.
    assert (desktop / "one" / "ledger.xlsx").read_bytes() == first
    assert (desktop / "two" / "ledger.xlsx").read_bytes() == second
