r"""Generalization harness for live planning. Requires a running Ollama.

Builds a HELD-OUT set that shares no filename, sheet name, label, domain or
phrasing with the supplied sample fixtures, then scores the live planner on it.
It exists to answer one question: does the planner generalize, or has the prompt
been tuned to the samples we happened to be given?

Every expectation is computed from the generated workbook's real layout, never
from a literal the planner could have copied out of the prompt.

Not part of the test suite: it needs a live model and is slow.

    & ".\.venv\Scripts\python.exe" tools\holdout_eval.py [runs]
"""

from __future__ import annotations

import shutil
import sys
import tempfile
from dataclasses import dataclass, field
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import fitz
from openpyxl import Workbook

from app.config import Settings
from app.planning.capabilities import detect_unsupported_request
from app.planning.exceptions import InvalidPlannerResponseError
from app.planning.planner import OllamaPlanner
from app.schemas import DraftPlan, TaskRequest


def make_pdf(path: Path, lines: list[str]) -> None:
    doc = fitz.open()
    page = doc.new_page()
    y = 72
    for line in lines:
        page.insert_text((72, y), line)
        y += 18
    doc.save(path)
    doc.close()


def make_workbook(
    path: Path,
    sheets: list[str],
    target_sheet: str,
    header_row: int,
    label_col: str,
    value_col: str,
    header: list[tuple[str, str]],
    labels: list[str],
) -> None:
    """Build a workbook with a deliberately non-default layout."""
    wb = Workbook()
    wb.remove(wb.active)
    for name in sheets:
        ws = wb.create_sheet(name)
        if name != target_sheet:
            ws["A1"] = "unrelated"
            continue
        for col, text in header:
            ws[f"{col}{header_row}"] = text
        for index, label in enumerate(labels):
            ws[f"{label_col}{header_row + 1 + index}"] = label
    wb.save(path)
    wb.close()


@dataclass
class Case:
    case_id: str
    prompt: str
    kind: str  # "write" | "refuse" | "clarify"
    source_name: str = ""
    workbook_name: str = ""
    sheet: str = ""
    cell: str = ""
    expected_value: float = 0.0
    notes: list[str] = field(default_factory=list)


def build_holdout(root: Path) -> list[Case]:
    cases: list[Case] = []

    # H1 — parts/units domain; target sheet is the SECOND sheet; labels in column
    # B; values in column D; header on row 3 so the first data row is row 4.
    make_pdf(
        root / "goods_received_note.pdf",
        [
            "Goods Received Note 4471",
            "SKU-88 quarter three units: 512",
            "SKU-91 quarter three units: 88",
            "SKU-104 quarter three units: 1340",
        ],
    )
    make_workbook(
        root / "warehouse_levels.xlsx",
        sheets=["Notes", "Warehouse", "Archive"],
        target_sheet="Warehouse",
        header_row=3,
        label_col="B",
        value_col="D",
        header=[("B", "Part"), ("C", "Bin"), ("D", "Q3 Units")],
        labels=["SKU-88", "SKU-91", "SKU-104"],
    )
    cases.append(
        Case(
            "H1-terse",
            "pull the Q3 units for SKU-88 out of goods_received_note.pdf "
            "into warehouse_levels.xlsx",
            "write",
            source_name="goods_received_note.pdf",
            workbook_name="warehouse_levels.xlsx",
            sheet="Warehouse",
            cell="D4",
            expected_value=512,
            notes=["2nd sheet", "labels col B", "values col D", "data starts row 4"],
        )
    )

    # H2 — headcount domain; target sheet is the THIRD sheet; labels column C;
    # values column E; header on row 2.
    make_pdf(
        root / "workforce_brief.pdf",
        [
            "Workforce Brief",
            "Engineering headcount: 214",
            "Finance headcount: 37",
            "Logistics headcount: 96",
        ],
    )
    make_workbook(
        root / "org_planning.xlsx",
        sheets=["Cover", "Budget", "Headcount"],
        target_sheet="Headcount",
        header_row=2,
        label_col="C",
        value_col="E",
        header=[("C", "Department"), ("D", "Site"), ("E", "People")],
        labels=["Engineering", "Finance", "Logistics"],
    )
    cases.append(
        Case(
            "H2-verbose",
            "I have a workforce brief saved as workforce_brief.pdf and a planning "
            "workbook called org_planning.xlsx. Please look up the Finance "
            "headcount in the brief and record it against Finance in the workbook.",
            "write",
            source_name="workforce_brief.pdf",
            workbook_name="org_planning.xlsx",
            sheet="Headcount",
            cell="E4",
            expected_value=37,
            notes=["3rd sheet", "labels col C", "values col E", "data starts row 3"],
        )
    )

    # H3 — ambiguity that is not a compass direction: two rows both match "Widget".
    make_pdf(
        root / "price_list.pdf",
        ["Price List", "Widget Standard unit price: 12.50", "Widget Pro unit price: 27.00"],
    )
    make_workbook(
        root / "catalogue.xlsx",
        sheets=["Pricing"],
        target_sheet="Pricing",
        header_row=1,
        label_col="A",
        value_col="B",
        header=[("A", "Product"), ("B", "Unit Price")],
        labels=["Widget Standard", "Widget Pro"],
    )
    cases.append(
        Case(
            "H3-ambiguous",
            "update the Widget unit price in catalogue.xlsx from price_list.pdf",
            "clarify",
            workbook_name="catalogue.xlsx",
            notes=["'Widget' matches two rows"],
        )
    )

    # H4 — genuinely unsupported by the action vocabulary.
    cases.append(
        Case(
            "H4-unsupported",
            "create a new PowerPoint deck summarising goods_received_note.pdf",
            "refuse",
        )
    )

    return cases


def score(case: Case, draft: DraftPlan) -> tuple[bool, list[str]]:
    problems: list[str] = []
    writes = [a for a in draft.actions if str(a.type) == "spreadsheet.write_cell"]

    if case.kind == "clarify":
        if writes:
            cells = [w.parameters.get("cell") for w in writes]
            problems.append(f"wrote to {cells} instead of asking which row was meant")
        return not problems, problems

    reads = [
        a
        for a in draft.actions
        if str(a.type).startswith(("pdf.", "document.", "file.read"))
    ]
    if not reads:
        problems.append("no source read step")
    elif case.source_name and not any(
        case.source_name in a.target for a in reads
    ):
        problems.append(f"source is {[a.target for a in reads]}, expected {case.source_name}")

    if len(writes) != 1:
        problems.append(f"expected 1 write, got {len(writes)}")
        return not problems, problems

    write = writes[0]
    if case.workbook_name and case.workbook_name not in write.target:
        problems.append(f"workbook is {write.target!r}")
    if write.parameters.get("sheet") != case.sheet:
        problems.append(
            f"sheet {write.parameters.get('sheet')!r} != actual {case.sheet!r}"
        )
    cell = write.parameters.get("cell")
    if not isinstance(cell, str) or cell.upper() != case.cell:
        problems.append(f"cell {cell!r} != actual {case.cell!r}")
    value = write.parameters.get("value")
    if not isinstance(value, dict) or "$ref" not in value:
        problems.append(f"value {value!r} is not $ref-bound (invented by the model)")
    return not problems, problems


def main() -> None:
    runs = int(sys.argv[1]) if len(sys.argv) > 1 else 1
    root = Path(tempfile.mkdtemp(prefix="ps2_holdout_"))
    print(f"held-out fixtures in {root}")
    cases = build_holdout(root)
    planner = OllamaPlanner(Settings())

    totals: dict[str, list[bool]] = {c.case_id: [] for c in cases}
    for run in range(1, runs + 1):
        print(f"\n{'=' * 78}\nRUN {run}\n{'=' * 78}")
        for case in cases:
            prompt = case.prompt.replace(
                "goods_received_note.pdf", str(root / "goods_received_note.pdf")
            )
            for name in (
                "warehouse_levels.xlsx",
                "workforce_brief.pdf",
                "org_planning.xlsx",
                "price_list.pdf",
                "catalogue.xlsx",
            ):
                prompt = prompt.replace(name, str(root / name))

            print(f"\n--- {case.case_id} ({case.kind}) {'/'.join(case.notes)}")
            refusal = detect_unsupported_request(prompt)
            if case.kind == "refuse":
                ok = refusal is not None
                totals[case.case_id].append(ok)
                print(f"    refusal: {refusal!r}")
                print(f"    -> {'PASS' if ok else 'FAIL'}")
                continue
            if refusal is not None:
                totals[case.case_id].append(False)
                print(f"    FAIL: wrongly refused ({refusal[:80]})")
                continue
            try:
                draft = planner._create_draft_sync(TaskRequest(text=prompt))
            except InvalidPlannerResponseError as exc:
                totals[case.case_id].append(False)
                print(f"    FAIL: {exc}")
                continue
            ok, problems = score(case, draft)
            totals[case.case_id].append(ok)
            for action in draft.actions:
                print(f"      {action.step_key}: {action.type} "
                      f"{Path(action.target).name!r} {action.parameters}")
            print(f"    -> {'PASS' if ok else 'FAIL'}")
            for problem in problems:
                print(f"       ! {problem}")

    print(f"\n{'=' * 78}\nHELD-OUT SUMMARY\n{'=' * 78}")
    for case in cases:
        results = totals[case.case_id]
        print(f"{case.case_id:18s} {sum(results)}/{len(results)}")
    flat = [r for results in totals.values() for r in results]
    print(f"{'OVERALL':18s} {sum(flat)}/{len(flat)}")
    shutil.rmtree(root, ignore_errors=True)


if __name__ == "__main__":
    main()
