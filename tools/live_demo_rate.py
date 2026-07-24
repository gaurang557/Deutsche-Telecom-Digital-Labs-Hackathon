r"""Five live runs of the sample demo sentence. Requires a running Ollama.

Scores against the workbook's ACTUAL layout, read from the file at run time, so
a plan that merely copied a cell out of the prompt cannot score a pass.

    & ".\.venv\Scripts\python.exe" tools\live_demo_rate.py
"""

from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import load_workbook

from app.config import Settings
from app.planning.exceptions import InvalidPlannerResponseError, PlannerUnavailableError
from app.planning.planner import OllamaPlanner
from app.schemas import TaskRequest

PACK = (
    Path.home()
    / "Desktop"
    / "ps2_mvp_test_fixtures"
    / "ps2_mvp_test_fixtures"
    / "fixtures"
)


def main() -> None:
    workbook = PACK / "results_blank.xlsx"
    wb = load_workbook(workbook)
    ws = wb["Summary"]
    rows = {
        ws.cell(r, 1).value: r
        for r in range(1, ws.max_row + 1)
        if ws.cell(r, 1).value
    }
    wb.close()
    expected_cell = f"B{rows['North']}"
    print(f"actual layout: {rows}  -> North's revenue cell is {expected_cell}")

    sentence = (
        f"in desktop fixtures Find the North Region revenue in "
        f"{PACK / 'quarterly_report.pdf'} and put it in the North row of {workbook}"
    )
    planner = OllamaPlanner(Settings())
    passes = 0
    for run in range(1, 6):
        try:
            draft = planner._create_draft_sync(TaskRequest(text=sentence))
        except (InvalidPlannerResponseError, PlannerUnavailableError) as exc:
            print(f"run {run}: REJECTED -> {exc}")
            continue
        writes = [a for a in draft.actions if str(a.type) == "spreadsheet.write_cell"]
        if not writes:
            print(f"run {run}: no write step; {[str(a.type) for a in draft.actions]}")
            continue
        params = writes[0].parameters
        value = params.get("value")
        bound = isinstance(value, dict) and "$ref" in value
        good = (
            params.get("sheet") == "Summary"
            and params.get("cell") == expected_cell
            and bound
        )
        passes += good
        print(
            f"run {run}: sheet={params.get('sheet')!r} cell={params.get('cell')!r} "
            f"value_bound_to_pdf={bound} -> {'PASS' if good else 'FAIL'}"
        )
    print(f"\nLIVE SUCCESS ON THE SAMPLE DEMO SENTENCE: {passes}/5")


if __name__ == "__main__":
    main()
